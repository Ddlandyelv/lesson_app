"""Excel 教案导出"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from .models import LessonData


def export_to_excel(data: LessonData, output_path: str):
    """将教案导出为格式化的 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "教案"

    # 样式定义
    title_font = Font(name="微软雅黑", size=16, bold=True)
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="微软雅黑", size=10)
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 列宽
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80

    row = 1

    # 标题行
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = f"《{data.title}》教案"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    row = 2

    # 信息行
    info_rows = [
        ("适用班级", data.level),
        ("课  时", data.duration),
        ("绘画材料", data.materials),
    ]
    for label, value in info_rows:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=row, column=2, value=value).font = body_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="center")
        row += 1

    # 分段内容
    sections = [
        ("教学目标", data.objectives),
        ("教学重点", data.key_points),
        ("教学难点", data.difficult_points),
        ("教学过程", data.process),
        ("课后延伸", data.extension),
        ("注意事项", data.notes),
    ]

    for section_title, content in sections:
        if not content.strip():
            continue

        ws.cell(row=row, column=1, value=section_title).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")

        ws.cell(row=row, column=2, value=content).font = body_font
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        # 根据内容行数自动调整行高
        line_count = content.count("\n") + 1
        ws.row_dimensions[row].height = max(30, line_count * 16)
        row += 1

    wb.save(output_path)
